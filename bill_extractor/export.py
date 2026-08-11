"""Native Excel and JSON exports."""

import json
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from bill_extractor.models import BillRecord
from bill_extractor.schema import FIELDS


def export_excel(records: list[BillRecord]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Extracted Bills"
    sheet.append([field.label for field in FIELDS])

    for record in records:
        sheet.append([record.values.get(field.key) if record.values.get(field.key) is not None else "-" for field in FIELDS])

    for column, field in enumerate(FIELDS, 1):
        if field.kind != "percent":
            continue
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, column)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00%"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, field in enumerate(FIELDS, 1):
        longest = max(len(field.label), *(len(str(sheet.cell(row, index).value)) for row in range(2, sheet.max_row + 1)))
        sheet.column_dimensions[get_column_letter(index)].width = min(max(longest + 2, 12), 34)

    details = workbook.create_sheet("Extraction Details")
    details.append(["Provider", "Filename", "Pages", "Warnings"])
    for record in records:
        details.append([record.provider, record.filename, ", ".join(map(str, record.pages)), " | ".join(record.warnings) or "-"])
    for cell in details[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    details.freeze_panes = "A2"
    details.column_dimensions["A"].width = 20
    details.column_dimensions["B"].width = 45
    details.column_dimensions["C"].width = 15
    details.column_dimensions["D"].width = 80

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def export_json(records: list[BillRecord]) -> str:
    return json.dumps([record.as_dict() for record in records], indent=2, ensure_ascii=False)
