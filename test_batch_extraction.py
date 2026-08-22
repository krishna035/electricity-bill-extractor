from pathlib import Path
from copy import deepcopy

from openpyxl import load_workbook
import pytest

from bill_extractor.export import export_excel
from bill_extractor.document import extract_pages
from bill_extractor.providers.ugvcl import UGVCLParser
from bill_extractor.presentation import display_dataframe
from bill_extractor.schema import FIELDS
from bill_extractor.service import InputFile, extract_files


DATA = Path("data")


def extract(filename: str):
    path = DATA / filename
    return extract_files([InputFile(path.name, path.read_bytes())], use_ocr=False)


def test_schema_matches_reference_workbook():
    workbook = load_workbook(DATA / "Angiplast_Bill_Analysis.xlsx", read_only=True)
    expected = [cell.value for cell in workbook["Angiplast"][1]][:58]

    fuel_percent_index = expected.index("Fuel Surcharge %") + 1
    expected[fuel_percent_index:fuel_percent_index] = ["Base FPPAS", "FPPAS Charges", "FPPAS %"]
    previous_dues_index = expected.index("Previous Dues") + 1
    expected[previous_dues_index:previous_dues_index] = [
        "Electricity Duty Credits",
        "TOU Charge Credits",
        "TDS Credits",
    ]
    expected.remove("Security Deposit Interest")
    expected.insert(expected.index("Other Credits"), "Security Deposit Interest")
    unit_rate_index = expected.index("Unit Rate (Total Consumption Charge)")
    demand_unit_rate_index = expected.index("Total Consumption-Demand Charge Unit Rate\n\n")
    expected[unit_rate_index], expected[demand_unit_rate_index] = (
        expected[demand_unit_rate_index],
        expected[unit_rate_index],
    )

    assert len(FIELDS) == 64
    assert expected == [field.label for field in FIELDS]


def test_ugvcl_merged_bill_matches_reference_readings():
    records = extract("Angiplast_UGVCL-INVOICES 2025-26 Merge.pdf")
    by_month = {record.values["billing_month"]: record for record in records}
    workbook = load_workbook(DATA / "Angiplast_Bill_Analysis.xlsx", data_only=True, read_only=True)
    sheet = workbook["Angiplast"]

    assert len(records) == 12
    for row in range(2, 14):
        month = sheet.cell(row, 1).value
        assert by_month[month].values["kwh_consumed"] == sheet.cell(row, 9).value
        assert by_month[month].values["total_consumption_charges"] == pytest.approx(sheet.cell(row, 38).value)

    april = by_month["APR-2025"].values
    assert april["customer_id"] == "65500"
    assert april["meter_number"] == "GHBD1806"
    assert april["solar_generation_units"] == 37920
    assert april["solar_setoff_units"] == 423
    assert april["solar_export_units"] == 5003
    assert april["solar_banking_units"] == 32917
    assert april["calculated_adjustment"] == 22648.10
    assert april["total_payable"] == 424939.03


def test_ugvcl_combines_multirow_adjustments():
    may = {record.values["billing_month"]: record for record in extract("Angiplast_UGVCL-INVOICES 2025-26 Merge.pdf")}["MAY-2025"].values

    assert may["solar_setoff_units"] == 2300
    assert may["solar_export_units"] == 18318
    assert may["solar_banking_units"] == 71442
    assert may["solar_generation_units"] == 89760
    assert may["security_deposit_interest"] == -115233.75
    assert may["other_credits"] == -975.24


def test_torrent_power_parser():
    record = extract("Torrent Bill Marking_260810_174038.pdf")[0]
    values = record.values

    assert record.provider == "Torrent Power"
    assert values["billing_month"] == "JAN-2026"
    assert values["customer_id"] == "100358213"
    assert values["contract_demand"] == 900
    assert values["actual_max_demand"] == 612
    assert values["billing_demand"] == 765
    assert values["kwh_consumed"] == 249690
    assert values["fuel_surcharge"] == 1003612.58
    assert values["base_fppas"] == 926797.08
    assert values["fppas_charges"] == 76815.50
    assert values["fppas_percent"] == pytest.approx(0.034)
    assert values["electricity_duty"] == 359765.81
    assert values["solar_generation_units"] == 37805
    assert values["solar_net_billed_units"] == 249139
    assert values["total_payable"] == 2779320.78
    assert not record.warnings


@pytest.mark.parametrize(
    ("filename", "amount_due"),
    [
        ("JAN 26.pdf", 2779320.78),
        ("FEB 26.pdf", 2858661.96),
    ],
)
def test_torrent_uses_precise_amount_due_and_excludes_demand_from_bd(filename, amount_due):
    record = extract(filename)[0]
    values = record.values

    assert values["total_payable"] == pytest.approx(amount_due)
    assert values["consumption_demand_unit_rate"] == pytest.approx(
        (values["total_consumption_charges"] - values["demand_charges"])
        / values["kwh_consumed"]
    )


def test_excel_export_uses_reference_headers_and_missing_marker():
    records = extract_files(
        [InputFile("S P METAL PGVCL.pdf", Path("S P METAL PGVCL.pdf").read_bytes())],
        use_ocr=False,
    )
    workbook = load_workbook(filename=__import__("io").BytesIO(export_excel(records)), data_only=True)
    sheet = workbook["Extracted Bills"]

    assert [cell.value for cell in sheet[1]] == [field.label for field in FIELDS]
    tariff_column = next(index for index, field in enumerate(FIELDS, 1) if field.key == "tariff_category")
    solar_column = next(index for index, field in enumerate(FIELDS, 1) if field.key == "solar_generation_units")
    assert sheet.cell(2, tariff_column).value == "LTMD"
    assert sheet.cell(2, solar_column).value == "-"


def test_percentage_columns_use_excel_percentage_format():
    records = extract("Angiplast_UGVCL-INVOICES 2025-26 Merge.pdf")
    workbook = load_workbook(filename=__import__("io").BytesIO(export_excel(records)), data_only=True)
    sheet = workbook["Extracted Bills"]
    solar_export_column = next(
        index for index, field in enumerate(FIELDS, 1) if field.key == "solar_export_percent"
    )

    assert sheet.cell(2, solar_export_column).value == pytest.approx(0.131935654)
    assert sheet.cell(2, solar_export_column).number_format == "0.00%"


def test_streamlit_preview_is_arrow_safe_with_mixed_missing_and_numeric_values():
    records = extract_files(
        [InputFile("S P METAL PGVCL.pdf", Path("S P METAL PGVCL.pdf").read_bytes())],
        use_ocr=False,
    )
    records.append(deepcopy(records[0]))
    records[1].values["kwh_increase_percent"] = 0.125

    frame = display_dataframe(records)

    assert all(str(dtype) == "string" for dtype in frame.dtypes)
    assert frame["% Increase in kWh"].tolist() == ["-", "0.125"]
    assert "% of Total units (Night)" in frame.columns
    assert "% of Total units (TOU)" in frame.columns
    assert frame.columns.is_unique

    pyarrow = pytest.importorskip("pyarrow")
    pyarrow.Table.from_pandas(frame)


def test_sparse_selectable_text_page_does_not_trigger_ocr():
    import fitz

    document = fitz.open()
    page = document.new_page()
    page.insert_text((72, 72), "This is a system generated bill. Hence no signature required.")
    data = document.tobytes()
    document.close()

    pages, warnings = extract_pages(data, "footer.pdf", use_ocr=True)

    assert pages[0].text.strip() == "This is a system generated bill. Hence no signature required."
    assert not pages[0].used_ocr
    assert not warnings


def test_ugvcl_wrapped_meter_and_adjustments_are_parsed():
    text = """
HT BILL FOR THE MONTH OF : APR-2026
Meter No: Make CTPT Make CTPT Srno CT Ratio PT Ratio Meter Status
Meter
Constant
GHBD0736 L&T 15 Normal
Adjustment Details Report for APR-2026/td>
Credit Board
39577.20 0.00 CREDIT BC FOR CONSUMPTION BETWEEN 11 AM TO 3 PM, MONTH-MAR-26, UNIT-65962
Charges
Credit Board
234.38 35.00 SOLAR ADJ FOR THE MONTH OF -MAR-26
Charges
Credit ED
4382.97 0.00 CREDIT ED FOR CONSUMPTION BETWEEN 11 AM TO 3 PM, MONTH-MAR-26, UNIT-65962
Charges
Credit ED
25.96 0.00 SOLAR ADJ FOR THE MONTH OF -MAR-26
Charges
Credit TDS 3202.00 0.00 TDS ADJUSTMENT 3-2026
Debit Banking
14016.00 0.00 SOLAR BANKING CHARGES FOR THE MONTH - MAR-26, UNIT-9344
Charge
"""

    values = UGVCLParser().parse(text)

    assert values["meter_number"] == "GHBD0736"
    assert values["solar_setoff_units"] == 35
    assert values["solar_setoff_credit"] == pytest.approx(-234.38)
    assert values["electricity_duty_credits"] == pytest.approx(-4408.93)
    assert values["tou_charge_credits"] == pytest.approx(-39577.20)
    assert values["tds_credits"] == pytest.approx(-3202.00)
    assert values["solar_banking_units"] == 9344
    assert values["solar_banking_charges"] == 14016
    assert values["other_credits"] == 0
    assert values["security_deposit_interest"] == 0


def test_ugvcl_other_debits_reconcile_adjustment():
    text = """
Adjustment Details Report for OCT-2025/td>
Credit Board Charges 848.74 127.00 SOLAR ADJ SEP-25
Credit ED Charges 93.77 0.00 SOLAR ADJ SEP-25
Credit TDS 3625.00 0.00 TDS ADJUSTMENT 9-2025
Debit Banking Charge 12012.00 0.00 SOLAR BANKING CHARGE SEP-25,U-8008
Debit Electricity Duty 0.05 0.00 ED RECOVERY AGAINST FC
Debit Fuel Surcharge 0.45 0.00 FC RECOVERY IN SOLAR SET OFF FOR THE MONTH OF JULY-25
"""

    values = UGVCLParser().parse(text)

    assert values["solar_setoff_units"] == 127
    assert values["solar_setoff_credit"] == pytest.approx(-848.74)
    assert values["electricity_duty_credits"] == pytest.approx(-93.77)
    assert values["tds_credits"] == pytest.approx(-3625.00)
    assert values["solar_banking_units"] == 8008
    assert values["other_credits"] == pytest.approx(0.50)


def test_photographed_bill_and_adjustment_are_merged():
    filenames = ["1000371676.jpg", "1000371677.jpg"]
    records = extract_files(
        [InputFile(name, (DATA / name).read_bytes()) for name in filenames],
        use_ocr=True,
    )

    assert len(records) == 1
    record = records[0]
    values = record.values
    assert record.provider == "UGVCL"
    assert values["billing_month"] == "MAY-2025"
    assert values["customer_id"] == "65500"
    assert values["tariff_category"] == "HTP-I"
    assert values["kwh_consumed"] == 52176
    assert values["solar_generation_units"] == 89760
    assert values["solar_setoff_units"] == 2300
    assert values["solar_export_units"] == 18318
    assert values["solar_banking_units"] == 71442
    assert values["net_payable"] == 295385.72
    assert values["total_payable"] == 295385.72
    assert not record.warnings


@pytest.mark.parametrize(
    (
        "filename",
        "setoff",
        "export",
        "banking",
        "banking_charge",
        "solar_credit",
        "setoff_credit",
        "electricity_duty_credit",
        "tds_credit",
        "other_credit",
    ),
    [
        ("1_APRIL-2025.pdf", 528, 202, 35318, 52977.00, -333.30, -3405.60, -510.84, -1108.00, 0.00),
        ("2_MAY-2025.pdf", 820, 1425, 53095, 79642.50, -2351.25, -5315.24, -1159.17, 0.00, -2412.54),
        ("3_JUNE-2025.pdf", 869, 1613, 51788, 77682.00, -2661.45, -5570.29, -835.54, 0.00, 0.00),
        ("4_JULY-2025.pdf", 1137, 719, 55333, 82999.50, -1186.35, -7219.95, -1082.99, -4177.00, 0.00),
        ("5_AUG-2025.pdf", 1731, 607, 27986, 41979.00, -1001.55, -10960.69, -1644.10, 0.00, 0.00),
    ],
)
def test_s21_ugvcl_adjustments(
    filename,
    setoff,
    export,
    banking,
    banking_charge,
    solar_credit,
    setoff_credit,
    electricity_duty_credit,
    tds_credit,
    other_credit,
):
    record = extract(filename)[0]
    values = record.values

    assert values["solar_setoff_units"] == setoff
    assert values["solar_export_units"] == export
    assert values["solar_banking_units"] == banking
    assert values["solar_generation_units"] == export + banking
    assert values["solar_net_billed_units"] == values["kwh_consumed"] - setoff
    assert values["solar_banking_charges"] == pytest.approx(banking_charge)
    assert values["solar_credit"] == pytest.approx(solar_credit)
    assert values["solar_setoff_credit"] == pytest.approx(setoff_credit)
    assert values["electricity_duty_credits"] == pytest.approx(electricity_duty_credit)
    assert values["tds_credits"] == pytest.approx(tds_credit)
    assert values["other_credits"] == pytest.approx(other_credit)
    assert values["calculated_adjustment"] == pytest.approx(values["advance_adjustment"])
    assert values["net_less_demand_unit_rate"] == pytest.approx(
        (values["total_payable"] - values["demand_charges"])
        / values["solar_net_billed_units"]
    )
    assert not record.warnings
